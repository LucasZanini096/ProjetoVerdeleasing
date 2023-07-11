# type: ignore 
# Selenium - Automatizando tarefas no navegador
from pathlib import Path
from time import sleep

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


# Caminho para a raiz do projeto
ROOT_FOLDER = Path(__file__).parent
# Caminho para a pasta onde o chromedriver está
CHROME_DRIVER_PATH = ROOT_FOLDER / 'drivers' / 'chromedriver'
EXCEL_FOLDER = "C:\\Users\\Administrador\\Downloads\\veiculos (5).xlsx"
EXCEL_FINAL = ROOT_FOLDER / 'workbook_final.xlsx'


def make_chrome_browser(*options: str) -> webdriver.Chrome:
    chrome_options = webdriver.ChromeOptions()

    # chrome_options.add_argument('--headless')
    if options is not None:
        for option in options:
            chrome_options.add_argument(option)

    chrome_service = Service(
        executable_path=str(CHROME_DRIVER_PATH),
    )

    browser = webdriver.Chrome(
        service=chrome_service,
        options=chrome_options
    )

    return browser


if __name__ == '__main__':
    
    if os.path.exists(EXCEL_FOLDER):
        os.remove(EXCEL_FOLDER)
    
    
    TIME_TO_WAIT = 10
        
    options = ()
    browser = make_chrome_browser(*options)

    browser.get("https://painel.cobli.co/#login")
    
    browser.maximize_window()

    SEARCH_INPUTS = ['username','password']
    INPUTS = ['ivan@verdeleasing.com.br', 'Sp18st21@']
        


    for index, search_input in enumerate(SEARCH_INPUTS):
        
        Input = WebDriverWait(browser, TIME_TO_WAIT).until(
         EC.presence_of_element_located(
                (By.NAME, search_input)
                )
            )
        Input.send_keys(INPUTS[index])
                
        
    Input.send_keys(Keys.ENTER)
    sleep(8)

    XPATHS = ['//*[@id="close"]', '//*[@id="container"]/div[1]/div[1]/div[4]/div[5]/a[1]/div[2]',
                '//*[@id="container"]/div[1]/div[1]/div[4]/div[5]/div[1]/div/div/a[1]/div[2]',
                '//*[@id="onesignal-slidedown-cancel-button"]',
                '//*[@id="content"]/div/div[2]/button/span']


    for Xpath in XPATHS:

        click = WebDriverWait(browser, TIME_TO_WAIT).until(
        EC.presence_of_element_located(
            (By.XPATH, Xpath)
         )
        )
    
        click.click()
        sleep(8)
            

    Base_Odometro_DF = pd.read_excel("C:\\Users\\Administrador\\Downloads\\veiculos (5).xlsx", engine="openpyxl")
    print(Base_Odometro_DF)
    Quilometragem_Veiculos = []
        
    for i in range(6, 14):
        Quilometragem_Veiculos.append(Base_Odometro_DF.iloc[i, 6])

    print(Quilometragem_Veiculos)

    workbook = Workbook()

    #Nome da planília
    sheet_name = "Planília Veículos"
    workbook.create_sheet(sheet_name, 0)
    worksheet: Worksheet = workbook.active

    for i in range(1,9):
        worksheet.cell(1,i,f'Veículo {str(i)}')
    
    worksheet.append(Quilometragem_Veiculos)

    workbook.save(EXCEL_FINAL)




    

    
        




